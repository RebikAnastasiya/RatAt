import os

from PyQt6 import (
    QtWidgets as qw,
    QtGui as qg,
    QtCore as qc
)

import mne, pandas as pd, numpy as np

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

import matplotlib
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg
from matplotlib.backends.backend_qtagg import NavigationToolbar2QT as NavigationToolbar
from matplotlib.backend_bases import key_press_handler as mpl_key_press_handler

from ratat_utils import *
from ratat_widgets import *

class PowerDensityWindow(qw.QMainWindow):

    is_power_plot = checkbox_accessor('is_power_cb')
    is_average_of_events = checkbox_accessor('is_average_over_events_cb')

    def closeEvent(self, event):
        # NOTE: this is hack
        plt.close()

    def is_channel_checked(self, ch_name):
        name = f'{ch_name}_cb'
        child = self.findChild(qc.QObject, name)
        if child is not None:
            val = child.property('checked')
        else:
            val = None

        return val

    def __init__(self, settings, main_filename, rawdata, on_close):
        super(PowerDensityWindow, self).__init__()

        self.Settings = settings
        self.MainFilename = main_filename
        self.rawdata = rawdata
        self.on_close = on_close

        cmap = matplotlib.cm.get_cmap("tab10")
        self.ch_colors = {
            ch_name: cmap(idx)
            for idx, ch_name in zip(range(len(rawdata.ch_names)), rawdata.ch_names)
        }

        self.init_window()
        self.init_elements()

        self.render_chart()

    def init_window(self):
        self.setWindowTitle('MNE Toolbox // Power Density Spectra')
        self.showMaximized()
        self.setWindowIcon(RatatIcon())

    def init_elements(self):
        main_layout = MainLayout()
        chart_layout = qw.QHBoxLayout()
        control_layout = qw.QVBoxLayout()
        control_layout.setContentsMargins(10,5,10,10)

        self.fig, self.ax = plt.subplots()

        self.canvas = FigureCanvasQTAgg(self.fig)
        self.canvas.mpl_connect('key_press_event', self.on_key_press)
        self.canvas.setFocusPolicy(qc.Qt.FocusPolicy.StrongFocus)

        self.toolbox = NavigationToolbar(self.canvas, self)
        main_layout.addWidget(self.toolbox)

        control_layout.addWidget(ToolboxHeader('Visible channels'))

        for ch in self.rawdata.ch_names:
            ch_checkbox_name = f"{ch}_cb"
            cb = qw.QCheckBox(ch, self)
            cb.setChecked(True)
            cb.stateChanged.connect(self.render_chart)
            cb.setObjectName(ch_checkbox_name)
            control_layout.addWidget(cb)

        control_layout.addWidget(ToolboxHeader('Chart options'))

        cb = qw.QCheckBox('Plot Power (not amplitude)', self)
        cb.setObjectName('is_power_cb')
        cb.setChecked(True)
        cb.stateChanged.connect(self.render_chart)
        control_layout.addWidget(cb)

        count_non_bad_events = count_non_bad_events_in_data(self.rawdata)
        if count_non_bad_events > 0:
            cb = qw.QCheckBox(f'Plot avarage of [{count_non_bad_events}] events', self)
            cb.setObjectName('is_average_over_events_cb')
            cb.stateChanged.connect(self.render_chart)
            control_layout.addWidget(cb)

        control_layout.addWidget(ToolboxHeader('Export'))

        btn = qw.QPushButton('Export to XLS', self)
        btn.clicked.connect(self.event__export)
        control_layout.addWidget(btn)

        control_layout.addStretch()

        main_layout.addLayout(chart_layout)
        chart_layout.addLayout(control_layout, stretch=1)
        chart_layout.addWidget(self.canvas, stretch=5)

        self.setCentralWidget(main_layout.get_widget())

    def render_chart(self):
        self.ax.clear()

        is_power_plot = self.is_power_plot
        ch_names = []

        for c in self.rawdata.ch_names:
            if self.is_channel_checked(c):
                ch_names.append(c)

        if len(ch_names) == 0:
            return

        if self.is_average_of_events:
            # NOTE: here create events from ALL NOT BAD annotations
            df = compute_average_psd_over_annotated_events(
                self.rawdata, ch_names
            )
            df.freq = df.index
        else:
            df = self.rawdata.compute_psd(
                method='welch',
                picks=ch_names,
                n_fft=2000
            ).to_data_frame()

        a_val = 0
        b_val = df.freq[np.where(df.freq > 5)[0][0]]
        idx_delta = np.logical_and(df.freq >= a_val, df.freq <= b_val)

        a_val = b_val
        b_val = df.freq[np.where(df.freq > 10)[0][0]]
        idx_delta_t = np.logical_and(df.freq >= a_val, df.freq <= b_val)

        a_val = b_val
        b_val = df.freq[np.where(df.freq > 20)[0][0]]
        idx_delta_b = np.logical_and(df.freq >= a_val, df.freq <= b_val)

        a_val = b_val
        b_val = df.freq[np.where(df.freq > 40)[0][0]]
        idx_delta_g = np.logical_and(df.freq >= a_val, df.freq <= b_val)

        fill_alpha = 0.05 * (len(self.rawdata.ch_names) - len(ch_names) + 1)

        ymax = 2e-9
        if not is_power_plot:
            ymax = np.sqrt(2e-9)
            for c_name in ch_names:
                df[c_name] = np.sqrt(df[c_name])

        for c_name in ch_names:
            alpha = fill_alpha
            self.ax.fill_between(df.freq, df[c_name], where=idx_delta, color='skyblue', alpha=alpha)
            self.ax.fill_between(df.freq, df[c_name], where=idx_delta_t, color='red', alpha=alpha)
            self.ax.fill_between(df.freq, df[c_name], where=idx_delta_b, color='green', alpha=alpha)
            self.ax.fill_between(df.freq, df[c_name], where=idx_delta_g, color='yellow', alpha=alpha)

        for c_name in ch_names:
            ch_color = self.ch_colors[c_name]
            self.ax.plot(df.freq, df[c_name], lw=2, label=c_name, color=ch_color)

        self.ax.set_xlabel('Frequency (Hz)')

        if is_power_plot:
            self.ax.set_ylabel('Power spectral density ($V^2 / Hz$)')
        else:
            self.ax.set_ylabel(r'Power spectral density ($V / \sqrt{Hz}$)')
        self.ax.set_xlim([0, 60])
        self.ax.set_ylim([0, ymax])
        self.ax.set_title("Welch's periodogram")

        self.ax.legend()

        self.canvas.draw()


    def on_key_press(self, event):
        mpl_key_press_handler(event, self.canvas, self.toolbox)

    @try_catch_btn
    def event__export(self, *args, **kwargs):
        report_file = get_step_filename(self.MainFilename, '02_psd', '__report', 'xlsx')

        wb = Workbook(report_file)

        df = self.rawdata.compute_psd(
            method='welch',
            n_fft=2000
        ).to_data_frame()

        # psd channels

        sheet_title = 'PSD channels'

        if not self.is_power_plot:
            sheet_title = 'PSD channels (Ampl)'
            for c_name in self.rawdata.ch_names:
                df[c_name] = np.sqrt(df[c_name])
            
        ws = wb.create_sheet(sheet_title)
        for r in dataframe_to_rows(df, index=True, header=True):
            ws.append(r)

        # freq range

        sheet_title = 'PSD average by group'

        mean_df = pd.DataFrame()
        mean_df['1 < f <= 5'] = df[(df.freq > 1) & (df.freq <= 5)].mean()
        mean_df['5 < f <= 10'] = df[(df.freq > 5) & (df.freq <= 10)].mean()
        mean_df['10 < f <= 20'] = df[(df.freq > 10) & (df.freq <= 20)].mean()
        mean_df['20 < f <= 50'] = df[(df.freq > 20) & (df.freq <= 50)].mean()

        if not self.is_power_plot:
            sheet_title = 'PSD average by group (Ampl)'
            for col_name in ['1 < f <= 5', '5 < f <= 10', '10 < f <= 20', '20 < f <= 50']:
                mean_df[col_name] = np.sqrt(mean_df[col_name])

        ws = wb.create_sheet(sheet_title)
        for r in dataframe_to_rows(mean_df, index=True, header=True):
            ws.append(r)
        del mean_df

        # average annotated events

        if count_non_bad_events_in_data(self.rawdata) > 0:
            # save averages
            sheet_title = 'PSD average in events'
            
            df = compute_average_psd_over_annotated_events(
                self.rawdata, self.rawdata.ch_names
            )

            if not self.is_power_plot:
                sheet_title = 'PSD average in events (Ampl)'
                for c_name in self.rawdata.ch_names:
                    df[c_name] = np.sqrt(df[c_name])

            ws = wb.create_sheet(sheet_title)
            for r in dataframe_to_rows(df, index=True, header=True):
                ws.append(r)

            # save freq range

            sheet_title = 'PSD average in events by group'

            df.freq = df.index

            mean_df = pd.DataFrame()
            mean_df['1 < f <= 5'] = df[(df.freq > 1) & (df.freq <= 5)].mean()
            mean_df['5 < f <= 10'] = df[(df.freq > 5) & (df.freq <= 10)].mean()
            mean_df['10 < f <= 20'] = df[(df.freq > 10) & (df.freq <= 20)].mean()
            mean_df['20 < f <= 50'] = df[(df.freq > 20) & (df.freq <= 50)].mean()

            if not self.is_power_plot:
                sheet_title = 'PSD average in events by group (Ampl)'
                for col_name in ['1 < f <= 5', '5 < f <= 10', '10 < f <= 20', '20 < f <= 50']:
                    mean_df[col_name] = np.sqrt(mean_df[col_name])

            ws = wb.create_sheet(sheet_title)
            for r in dataframe_to_rows(mean_df, index=True, header=True):
                ws.append(r)
            
            del mean_df
            del df

        wb.save(report_file)
        msg_success(f'Saved file: {report_file}')
        log_i('PSD reprot saved [{0}]', report_file)



def count_non_bad_events_in_data(rawdata):
    '''count events in that are not classified as BAD_*'''
    count_non_bad_events = 0
    for idx, onset in enumerate(rawdata.annotations.onset):
        description = rawdata.annotations.description[idx]

        if description.startswith('BAD_'):
            continue

        count_non_bad_events += 1

    return count_non_bad_events


def compute_average_psd_over_annotated_events(rawdata, ch_names=None):
    '''compute avarage PSD for specified channells'''
    events, event_ids = mne.events_from_annotations(rawdata)
    epochs = mne.Epochs(rawdata, events, event_id=event_ids)
    df = epochs.compute_psd(
        method='welch',
        picks=ch_names
    ).to_data_frame()

    ch_avg = df.pivot_table(
        values=ch_names,
        index = ['freq'],
        aggfunc = 'mean'
    )
    return ch_avg

def compute_and_save_in_excel(rawdata, wb):
    '''generate psd xlsx report.
    created sheets
    - PSD means - average PSD for specific frequency ranges
    - PSD raw data - computed raw PSD data from MNE
    '''
    df = rawdata.compute_psd(n_fft=2000).to_data_frame()

    mean_df = pd.DataFrame()
    mean_df['1 < f <= 5'] = df[(df.freq > 1) & (df.freq <= 5)].mean()
    mean_df['5 < f <= 10'] = df[(df.freq > 5) & (df.freq <= 10)].mean()
    mean_df['10 < f <= 20'] = df[(df.freq > 10) & (df.freq <= 20)].mean()
    mean_df['20 < f <= 50'] = df[(df.freq > 20) & (df.freq <= 50)].mean()

    mean_df.drop(['freq'], inplace=True)

    ws = wb.create_sheet('PSD means')
    for r in dataframe_to_rows(mean_df, index=True, header=True):
        ws.append(r)

    ws = wb.create_sheet('PSD raw data')
    for r in dataframe_to_rows(df, index=True, header=True):
        ws.append(r)


if __name__ == '__main__':
    import sys
    from PyQt6.QtWidgets import QApplication

    init_application()

    abs_filename = PARSED_ARGS['file']

    assert abs_filename is not None, '--file argument has to be specified'

    app = QApplication(sys.argv)
    s = SettingState()
    s.load_from_json()

    rawdata = read_mne_rawdata(abs_filename)

    w = PowerDensityWindow(s, abs_filename, rawdata, None)

    app_run_main_window(app, w)
